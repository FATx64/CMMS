from __future__ import annotations

import datetime
import random
import uuid
from io import BytesIO
from pathlib import Path
from typing import TYPE_CHECKING, Any

from django.core.files.base import File
from django.core.files.uploadedfile import UploadedFile
from django.forms.utils import flatatt
from django.forms.widgets import static
from django.http.request import HttpRequest
from django.utils.html import format_html, html_safe, mark_safe
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image

from cmms import models


CMMS_EPOCH: int = 1672531200000


def utcnow() -> datetime.datetime:
    return datetime.datetime.now(datetime.timezone.utc)


def snowflake(dt: datetime.datetime | None = None, /, *, high: bool = False) -> int:
    """Simplified version of how discord ID generated

    REF: https://github.com/Rapptz/discord.py/blob/e870bb1335e3f824c83a40df4ea9b17f215fde63/discord/utils.py#L395-L422
    """
    if not dt:
        dt = utcnow()
    millis = int(dt.timestamp() * 1000 - CMMS_EPOCH)
    return (millis << 22) + (2**22 - 1 if high else 0) + random.randint(0, 10)


def time_from_snowflake(id: int, /) -> datetime.datetime:
    """REF: https://github.com/Rapptz/discord.py/blob/e870bb1335e3f824c83a40df4ea9b17f215fde63/discord/utils.py#L375-L392"""
    millis = ((id >> 22) + CMMS_EPOCH) / 1000
    return datetime.datetime.fromtimestamp(millis, tz=datetime.timezone.utc)


def generate_hexa_id() -> str:
    """Generate 32 long hexa id"""
    return uuid.uuid1().hex


def handle_image_upload(path: Path, file: UploadedFile) -> str:
    _id = generate_hexa_id()
    path.mkdir(parents=True, exist_ok=True)

    buffer = BytesIO()
    for chunk in File(file):
        buffer.write(chunk)

    buffer.seek(0)
    img = Image.open(buffer)
    img.save(path / f"{_id}.webp", "WEBP")
    return _id


def handle_avatar_upload(user_id: int, file: UploadedFile) -> str:
    return handle_image_upload(Path(f"data/avatars/{user_id}"), file)


def handle_equipment_pict_upload(equipment_id: int, file: UploadedFile) -> str:
    return handle_image_upload(Path(f"data/pictures/equipment/{equipment_id}"), file)


def handle_sparepart_pict_upload(sparepart_id: int, file: UploadedFile) -> str:
    return handle_image_upload(Path(f"data/pictures/sparepart/{sparepart_id}"), file)


@html_safe
class JS:
    def __init__(self, js: str | None, attrs: dict[str, Any] | None = None):
        self.js = js
        self.attrs = attrs or {}

    def __str__(self):
        if self.js is None:
            return format_html("<script {}></script>", mark_safe(flatatt(self.attrs)))
        return format_html(
            '<script src="{}"{}></script>',
            self.js if self.js.startswith(("http://", "https://", "/")) else static(self.js),
            mark_safe(flatatt(self.attrs)),
        )


class CMMSRequest(HttpRequest):
    if TYPE_CHECKING:
        user: models.User


def number_to_letter(number):
    """
    Converts a number to the corresponding letter in Excel column notation.
    Example: 1 -> 'A', 2 -> 'B', 27 -> 'AA', 28 -> 'AB', etc.
    """
    letters = ""
    while number > 0:
        number, remainder = divmod(number - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def export_workorder_to_xl():
    wb = Workbook()
    ws = wb.active
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 32
    data = models.WorkOrder.objects.all()

    row_height = 32
    ws.row_dimensions[1].height = row_height
    title = ["Code", "Description", "Start Date", "End Date", "Cost", "Work Center", "Location", "Equipment"]
    ws.append(title)
    ft = Font(color="FFFFFF")
    fill = PatternFill("solid", fgColor="75DD79")
    side = Side("medium", "000000")
    border = Border(left=side, right=side, top=side, bottom=side)
    align = Alignment("center", "center", wrapText=True)
    for row in ws[f"A1:{number_to_letter(ws.max_column)}1"]:
        for cell in row:
            cell.font = ft
            cell.fill = fill
            cell.border = border
            cell.alignment = align

    to_be_added = []
    for wo in data:
        to_be_added.append(
            [
                f"{wo.type}-{wo.code}",
                f"{wo.type} - {wo.description}",
                wo.start_date,
                wo.end_date,
                wo.equipment.cost,
                wo.equipment.work_place.name,
                wo.equipment.location,
                wo.equipment.work_order_format(),
            ]
        )
    for i in to_be_added:
        ws.append(i)
        ws.row_dimensions[ws.max_row].height = row_height
        for row in ws[f"A{ws.max_row}:{number_to_letter(ws.max_column)}{ws.max_row}"]:
            for cell in row:
                cell.border = border
                cell.alignment = align
        ws[f"E{ws.max_row}"].number_format = '"Rp"#,##0.00'

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
